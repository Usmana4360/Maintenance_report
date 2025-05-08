import streamlit as st
import pandas as pd
import requests
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment

# Constants
EXCEL_FILE = "generated_reports.xlsx"
HF_API_KEY = os.getenv("HF_API_KEY")  # Add this to your .env or Streamlit secrets
HF_API_URL = "https://api-inference.huggingface.co/models/mistralai/Mistral-7B-Instruct-v0.1"

# Generate report using Mistral via Hugging Face
def generate_report(unit, machine, technician_name, issue):
    """Generates a professional maintenance report using Mistral. Falls back to raw issue if error occurs."""
    prompt = f"""
You are an expert electrical maintenance engineer. Generate a **concise and professional** one-line report based on the following details:

Unit: {unit}
Machine: {machine}
Technician Name: {technician_name}
Issue Reported: {issue}

Instructions:
- Maintain a clear, formal, technical tone.
- Mention unit, machine, technician, and issue.
- Keep it one line only.

Example:
"Technician {technician_name} diagnosed {issue} on {machine} in {unit}, performed necessary maintenance, and restored functionality."

Now generate the report.
    """

    headers = {
        "Authorization": f"Bearer {HF_API_KEY}",
        "Content-Type": "application/json"
    }

    payload = {
        "inputs": prompt,
        "parameters": {
            "max_new_tokens": 100,
            "temperature": 0.7,
            "return_full_text": False
        }
    }

    try:
        response = requests.post(HF_API_URL, headers=headers, json=payload)
        if response.status_code == 200:
            result = response.json()
            return result[0]["generated_text"].strip()
        else:
            print(f"⚠ Mistral API error {response.status_code}: {response.text}")
            return f"Issue reported and solved by {technician_name}: {issue}"
    except Exception as e:
        print(f"❌ Exception during API call: {str(e)}")
        return f"Issue reported and solved by {technician_name}: {issue}"


# Save report to Excel
def save_to_excel(unit, machine, technician_name, issue, report):
    columns = ["Date", "Unit", "Machine", "Technician Name", "Issue", "Generated Report"]

    new_data = pd.DataFrame([{
        "Date": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Unit": unit,
        "Machine": machine,
        "Technician Name": technician_name,
        "Issue": issue,
        "Generated Report": report
    }])

    if os.path.exists(EXCEL_FILE):
        try:
            existing_data = pd.read_excel(EXCEL_FILE, engine="openpyxl")
            updated_data = pd.concat([existing_data, new_data], ignore_index=True)
        except Exception as e:
            print(f"Error reading Excel: {e}")
            updated_data = new_data
    else:
        updated_data = new_data

    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="w") as writer:
        updated_data.to_excel(writer, index=False, sheet_name="Work Orders")

    # Format Excel
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="center")
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    for cell in ws[1]:
        cell.font = Font(bold=True)
    wb.save(EXCEL_FILE)
    print("✅ Report saved and formatted.")

# Load existing reports
def load_reports():
    if os.path.exists(EXCEL_FILE):
        return pd.read_excel(EXCEL_FILE)
    return pd.DataFrame(columns=["Unit", "Machine", "Technician", "Issue", "Report"])

# Streamlit UI
st.set_page_config(page_title="AI Maintenance Report Generator", layout="wide")
st.title("🔧 AI-Powered Maintenance Report Generator (Mistral)")

col1, col2 = st.columns(2)
with col1:
    unit = st.text_input("Enter Unit:", placeholder="e.g., Unit 5")
    technician_name = st.text_input("Enter Technician Name:", placeholder="e.g., John Doe")
with col2:
    machine = st.text_input("Enter Machine Name:", placeholder="e.g., Compressor A1")
    issue = st.text_input("Enter Issue Reported:", placeholder="e.g., High temperature issue")

if st.button("Generate Report"):
    if unit and machine and technician_name and issue:
        with st.spinner("Generating report using Mistral..."):
            report = generate_report(unit, machine, technician_name, issue)
            save_to_excel(unit, machine, technician_name, issue, report)
            st.success("✅ Report generated and saved!")
            st.subheader("Generated Report:")
            st.write(report)
    else:
        st.warning("⚠ Please fill in all fields before generating the report.")

# Report history
st.subheader("📁 Report History")
df_reports = load_reports()
st.dataframe(df_reports, use_container_width=True)

if os.path.exists(EXCEL_FILE):
    with open(EXCEL_FILE, "rb") as file:
        st.download_button("📥 Download Reports", file, file_name="generated_reports.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
